from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_POST, require_GET
from matricula.scripts.cargar_tree import crear_datos_prueba
from matricula.scripts.cargar_tree_apre import crear_datos_prueba_aprendiz
from django.db import transaction
from django.contrib.staticfiles import finders
from django.contrib.contenttypes.models import ContentType
import io
import zipfile
import os
import openpyxl
import csv
import random
import string
import json
from commons.permisos import bloquear_si_consulta
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.db.models import Q, Count, Value
from django.db.models.functions import Coalesce
import logging
from django.http import FileResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from .forms import EncuApreForm, EncuentroForm, ProgramaForm, RapsForm
from commons.models import T_encu, AuditLog, T_departa, T_munici, T_gestor_depa, T_fase, T_centro_forma, T_prematri_docu, T_docu, T_munici, T_insti_edu, T_perfil, T_DocumentFolderAprendiz, T_encu_apre, T_apre, T_ficha, T_progra, T_fase_ficha, T_instru, T_perfil, T_compe, T_raps, T_DocumentFolder, T_repre_legal, T_grupo, T_gestor, T_gestor_grupo
from django.contrib.auth.decorators import login_required
from datetime import datetime
from django.conf import settings
from django.core.files.storage import default_storage
from django.contrib.auth.models import User
from io import TextIOWrapper
from django.forms import ValidationError
from django.core.validators import validate_email

logger = logging.getLogger('django')

###############################################################################################################
#        VISTAS FICHA
###############################################################################################################

@login_required
def fichas(request):
    return render(request, 'listar_fichas.html')


@require_GET
@login_required
def obtener_opciones_fichas_estados(request):
    perfil_logueado = T_perfil.objects.get(user=request.user)

    fichas = T_ficha.objects.filter(esta__isnull=False)

    if perfil_logueado.rol == "gestor":
        gestor = T_gestor.objects.get(perfil=perfil_logueado)
        departamentos = T_gestor_depa.objects.filter(
            gestor=gestor).values_list('depa__nom_departa', flat=True)
        fichas = fichas.filter(
            insti__muni__nom_departa__nom_departa__in=departamentos)

    estados = fichas.distinct().values_list('esta', flat=True)
    return JsonResponse(list(estados), safe=False)


@require_GET
@login_required
def obtener_opciones_fichas_instructores(request):
    perfil_logueado = T_perfil.objects.get(user=request.user)

    if perfil_logueado.rol == "instructor":
        return JsonResponse([perfil_logueado.nom], safe=False)

    fichas = T_ficha.objects.filter(instru__isnull=False)

    if perfil_logueado.rol == "gestor":
        gestor = T_gestor.objects.get(perfil=perfil_logueado)
        departamentos = T_gestor_depa.objects.filter(
            gestor=gestor).values_list('depa__nom_departa', flat=True)
        fichas = fichas.filter(
            insti__muni__nom_departa__nom_departa__in=departamentos)

    instructores = fichas.distinct().values_list('instru__perfil__nom', flat=True)
    opciones = ['Sin asignar'] + list(instructores)
    return JsonResponse(opciones, safe=False)


@require_GET
@login_required
def obtener_opciones_fichas_programas(request):
    perfil_logueado = T_perfil.objects.get(user=request.user)

    fichas = T_ficha.objects.filter(progra__isnull=False)

    if perfil_logueado.rol == "gestor":
        gestor = T_gestor.objects.get(perfil=perfil_logueado)
        departamentos = T_gestor_depa.objects.filter(
            gestor=gestor).values_list('depa__nom_departa', flat=True)
        fichas = fichas.filter(
            insti__muni__nom_departa__nom_departa__in=departamentos)

    programas = fichas.distinct().values_list('progra__nom', flat=True)
    return JsonResponse(list(programas), safe=False)


@require_GET
@login_required
def filtrar_fichas(request):
    perfil_logueado = T_perfil.objects.get(user=request.user)
    estados = request.GET.getlist('estados', [])
    instructores = request.GET.getlist('instructores', [])
    programas = request.GET.getlist('programas', [])

    fichas = T_ficha.objects.all()

    if perfil_logueado.rol == "instructor":
        instructor = T_instru.objects.get(perfil=perfil_logueado)
        fichas = fichas.filter(instru=instructor)
    elif perfil_logueado.rol == "gestor":
        gestor = T_gestor.objects.get(perfil=perfil_logueado)
        departamentos = T_gestor_depa.objects.filter(
            gestor=gestor).values_list('depa__nom_departa', flat=True)
        fichas = fichas.filter(
            insti__muni__nom_departa__nom_departa__in=departamentos)

    if estados:
        fichas = fichas.filter(esta__in=estados)
    if instructores:
        incluir_null = 'Sin asignar' in instructores
        nombres_validos = [i for i in instructores if i != 'Sin asignar']

        if incluir_null and nombres_validos:
            fichas = fichas.filter(
                Q(instru__perfil__nom__in=nombres_validos) |
                Q(instru__isnull=True)
            )
        elif incluir_null:
            fichas = fichas.filter(instru__isnull=True)
        else:
            fichas = fichas.filter(instru__perfil__nom__in=nombres_validos)

    if programas:
        fichas = fichas.filter(progra__nom__in=programas)

    data = [
        {
            'id': f.id,
            'num': f.num,
            'grupo': f.grupo.id,
            'estado': f.esta,
            'fecha_aper': f.fecha_aper.strftime('%d/%m/%Y') if f.fecha_aper else None,
            'fecha_cierre': f.fecha_cierre.strftime('%d/%m/%Y') if f.fecha_cierre else None,
            'centro': f.centro.nom,
            'institucion': f.insti.nom,
            'instru': f.instru.perfil.nom if f.instru else None,
            'matricu': f.num_apre_proce,
            'progra': f.progra.nom
        } for f in fichas
    ]
    return JsonResponse(data, safe=False)


@require_POST
@login_required
@bloquear_si_consulta
def cambiar_numero_ficha(request, ficha_id):
    nuevo_num = request.POST.get('nuevo_num', '').strip()

    # Validación básica: campo vacío
    if not nuevo_num:
        return JsonResponse({'status': 'error', 'message': 'Número de ficha inválido.'}, status=400)

    # Validación: solo números
    if not nuevo_num.isdigit():
        return JsonResponse({'status': 'error', 'message': 'El número de ficha debe contener solo dígitos.'}, status=400)

    # Validación: número duplicado (excluyendo el actual)
    if T_ficha.objects.filter(num=nuevo_num).exclude(id=ficha_id).exists():
        return JsonResponse({'status': 'error', 'message': 'Ya existe una ficha con ese número.'}, status=400)

    ficha = get_object_or_404(T_ficha, id=ficha_id)
    ficha.num = nuevo_num
    ficha.save()

    return JsonResponse({'status': 'success', 'message': 'Número de ficha actualizado correctamente.'}, status=200)


@login_required
def listar_fichas(request):
    perfil = getattr(request.user, 't_perfil', None)
    if perfil is None or perfil.rol != 'instructor':
        return render(request, 'fichas.html', {'mensaje': 'No tienes permisos para acceder a esta página.'})

    instructor = T_instru.objects.filter(perfil=perfil).first()

    if not instructor:
        return render(request, 'fichas.html', {'mensaje': 'No se encontraron fichas asociadas a este instructor.'})

    fichas = T_ficha.objects.filter(instru=instructor)

    print(fichas)
    return render(request, 'fichas.html', {'fichas': fichas})


@login_required
def panel_ficha(request, ficha_id):
    ficha = get_object_or_404(T_ficha, id=ficha_id)
    fase = T_fase_ficha.objects.filter(ficha_id=ficha_id, vige=1).first()
    aprendices = T_apre.objects.filter(ficha=ficha_id)

    return render(request, 'panel_ficha.html', {
        'ficha': ficha,
        'fase': fase,
        'aprendices': aprendices,
    })


@login_required
def obtener_carpetas(request, ficha_id):
    # Obtener todas las carpetas y documentos asociados a la ficha
    nodos = T_DocumentFolder.objects.filter(ficha_id=ficha_id).values(
        "id", "name", "parent_id", "tipo", "documento__id", "documento__nom", "documento__archi"
    )

    folder_map = {}

    for nodo in nodos:
        nodo_id = nodo["id"]
        parent_id = nodo["parent_id"]

        nodo_data = {
            "id": nodo_id,
            "name": nodo["name"],
            "parent_id": parent_id,
            "tipo": nodo["tipo"],
            "children": []
        }

        if nodo["tipo"] == "documento":
            nodo_data.update({
                "documento_id": nodo["documento__id"],
                "documento_nombre": nodo["documento__nom"],
                "url": nodo["documento__archi"],  # La URL del archivo
            })

        folder_map[nodo_id] = nodo_data

    root_nodes = []

    for nodo in folder_map.values():
        parent_id = nodo["parent_id"]
        if parent_id:
            if parent_id in folder_map:
                folder_map[parent_id]["children"].append(nodo)
        else:
            root_nodes.append(nodo)

    return JsonResponse(root_nodes, safe=False)


@login_required
def descargar_portafolio_zip(request, ficha_id):
    nodos = T_DocumentFolder.objects.filter(
        ficha_id=ficha_id).select_related("documento")

    # Creamos un diccionario para construir la jerarquía de carpetas
    folder_map = {}

    for nodo in nodos:
        folder_map[nodo.id] = {
            "id": nodo.id,
            "name": nodo.name,
            "parent_id": nodo.parent_id,
            "tipo": nodo.tipo,
            "documento": nodo.documento if nodo.tipo == "documento" else None,
            "children": []
        }

    # Construimos la jerarquía
    root_nodes = []
    for nodo in folder_map.values():
        if nodo["parent_id"]:
            parent = folder_map.get(nodo["parent_id"])
            if parent:
                parent["children"].append(nodo)
        else:
            root_nodes.append(nodo)

    def agregar_a_zip(zip_file, nodo, ruta_actual):
        nombre = nodo["name"]
        ruta = os.path.join(ruta_actual, nombre)

        if nodo["tipo"] == "carpeta":
            # Creamos una carpeta vacía en el zip
            zip_file.writestr(f"{ruta}/", "")
            for hijo in nodo["children"]:
                agregar_a_zip(zip_file, hijo, ruta)
        elif nodo["tipo"] == "documento" and nodo["documento"] and nodo["documento"].archi:
            path_archivo = nodo["documento"].archi.path
            try:
                with open(path_archivo, "rb") as f:
                    contenido = f.read()
                    zip_file.writestr(f"{ruta}", contenido)
            except FileNotFoundError:
                pass
    # Creamos el zip en memoria
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for nodo in root_nodes:
            agregar_a_zip(zip_file, nodo, "")

    buffer.seek(0)
    
    AuditLog.objects.create(
        user=request.user,
        action="download",
        content_type=ContentType.objects.get_for_model(T_DocumentFolder),
        object_id=None,
        related_id=ficha_id,
        related_type="ficha",
        extra_data=f"Descargó el portafolio completo de la ficha {ficha_id} "
                   f"con {nodos.count()} nodos (documentos y carpetas)."
    )

    response = HttpResponse(buffer, content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename=portafolio_ficha_{ficha_id}.zip'
    return response


@login_required
def cargar_documento(request):
    if request.method == 'POST':
        if not request.FILES.get("file"):
            return JsonResponse({'status': 'error', 'message': 'Debe cargar un documento'}, status=400)

        file = request.FILES["file"]
        
        if file.size == 0:
            return JsonResponse({'status': 'error', 'message': 'El archivo está vacío o la carga falló'}, status=400)

        folder_id = request.POST.get("folder_id")
        contexto = request.POST.get("contexto")

        MODEL_MAP = {
            "ficha": T_DocumentFolder,
            "aprendiz": T_DocumentFolderAprendiz
        }

        if contexto not in MODEL_MAP:
            return JsonResponse({'status': 'error', 'message': 'Contexto no válido'}, status=400)

        model = MODEL_MAP[contexto]
        folder = get_object_or_404(model, id=folder_id)

        # Construir ruta dinámica
        if contexto == "ficha":
            ruta = f'documentos/fichas/portafolio/{folder.ficha.id}/{file.name}'
        else:  # aprendiz
            ruta = f'documentos/fichas/portafolio/aprendices/{folder.aprendiz.id}/{file.name}'

        # Guardar el archivo
        ruta_guardada = default_storage.save(ruta, file)

        size_kb = max(1, file.size // 1024) if file.size >= 1024 else f"{file.size} B"


        # Crear registro en T_docu
        new_docu = T_docu.objects.create(
            nom=file.name,
            tipo=file.name.split('.')[-1],
            tama=size_kb,
            archi=ruta_guardada,
            priva='No',
            esta='Activo'
        )

        # Crear el nodo del árbol con kwargs según contexto
        if contexto == "ficha":
            extra_kwargs = {"ficha": folder.ficha}
        else:  # aprendiz
            extra_kwargs = {"aprendiz": folder.aprendiz}

        with transaction.atomic():
            document_node = model.objects.create(
                name=file.name,
                parent=folder,
                tipo="documento",
                documento=new_docu,
                **extra_kwargs
            )
            
            if contexto == "ficha":
                related_id = document_node.ficha_id
            elif contexto == "aprendiz":
                related_id = document_node.aprendiz_id
            else:
                related_id = None
                
            extra_data = (
                f"Se cargó el documento '{document_node.name}' "
                f"en la carpeta {folder.id} "
            )
            
            AuditLog.objects.create(
              user= request.user,
              action="create",
              content_type= ContentType.objects.get_for_model(document_node),
              object_id= document_node.id,
              related_id = related_id,
              related_type = contexto,
              extra_data=extra_data
            )


        return JsonResponse({
            'status': 'success',
            'message': 'Documento cargado con éxito',
            'document': {
                'id': document_node.id,
                'name': document_node.name,
                'url': new_docu.archi.url,
                'folder_id': folder.id,
                'tipo': 'documento'
            }
        }, status=200)

    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)


@login_required
def mover_documento(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            document_id = data.get("document_id")
            target_folder_id = data.get("target_folder_id")
            contexto = data.get("contexto")

            if not document_id or not target_folder_id or not contexto:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Faltan parámetros: document_id o target_folder_id o contexto'
                }, status=400)

            MODEL_MAP = {
                "ficha": T_DocumentFolder,
                "aprendiz": T_DocumentFolderAprendiz
            }

            if contexto not in MODEL_MAP:
                return JsonResponse({'status': 'error', 'message': 'Contexto no válido'}, status=400)

            model = MODEL_MAP[contexto]

            # Obtener el documento y la carpeta destino
            documento_node = get_object_or_404(
                model, id=document_id, tipo="documento")
            carpeta_destino = get_object_or_404(
                model, id=target_folder_id, tipo="carpeta")

            # Evitar mover dentro de sí mismo o a un documento
            if documento_node.id == carpeta_destino.id:
                return JsonResponse({
                    'status': 'error',
                    'message': 'No se puede mover un documento dentro de sí mismo.'
                }, status=400)

            carpeta_origen = documento_node.parent  

            # Actualizar la carpeta padre
            documento_node.parent = carpeta_destino
            documento_node.save()

            # Registrar en el log
            if contexto == "ficha":
                related_id = documento_node.ficha_id
            elif contexto == "aprendiz":
                related_id = documento_node.aprendiz_id
            else:
                related_id = None

            AuditLog.objects.create(
                user=request.user,
                action="update",
                content_object=documento_node,
                related_id = related_id,
                related_type = contexto,
                extra_data=(
                    f"Se movió el documento '{documento_node.name}' "
                    f"de la carpeta {carpeta_origen.id} a la carpeta {carpeta_destino.id} "
                )
            )

            return JsonResponse({
                'status': 'success',
                'message': 'Documento movido con éxito',
                'document': {
                    'id': documento_node.id,
                    'name': documento_node.name,
                    'folder_id': carpeta_destino.id
                }
            }, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Formato JSON inválido'}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)


@login_required
def eliminar_documento_portafolio_ficha(request, documento_id):
    if request.method != "DELETE":
        return JsonResponse({"status": "error", "error": "Método no permitido"}, status=405)

    documento = get_object_or_404(T_DocumentFolder, id=documento_id)

    nombre_doc = documento.name if (documento.name) else "Documento sin archivo"
    extra_data = (
        f"Se elimino el documento '{nombre_doc}' "
        f"en la carpeta {documento.parent_id} "
    )
    
    # Si hay un archivo relacionado, eliminarlo manualmente
    if documento.documento and documento.documento.archi:
        archivo_a_eliminar = documento.documento.archi.name
        if archivo_a_eliminar:
            default_storage.delete(archivo_a_eliminar)
        documento.documento.delete()

    documento.delete()
    
    AuditLog.objects.create(
        user=request.user,
        content_type=ContentType.objects.get_for_model(documento),
        object_id=documento.id,
        action="delete",
        related_id = documento.ficha_id,
        related_type = "ficha",
        extra_data=extra_data
    )

    return JsonResponse({"status": "success", "message": "Eliminado correctamente"}, status=200)


@login_required
def obtener_hijos_carpeta(request, carpeta_id):
    try:
        nodos = T_DocumentFolder.objects.filter(parent_id=carpeta_id).values(
            "id", "name", "parent_id", "tipo", "documento__id", "documento__nom", "documento__archi"
        )

        children = []

        for nodo in nodos:
            data = {
                "id": nodo["id"],
                "name": nodo["name"],
                "parent_id": nodo["parent_id"],
                "tipo": nodo["tipo"],
                "children": []
            }

            if nodo["tipo"] == "documento":
                data.update({
                    "documento_id": nodo["documento__id"],
                    "documento_nombre": nodo["documento__nom"],
                    "url": nodo["documento__archi"],
                })

            children.append(data)

        return JsonResponse(children, safe=False)

    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


@login_required
def obtener_carpetas_aprendiz(request, aprendiz_id):
    # Obtener todas las carpetas y documentos asociados a la ficha
    nodos = T_DocumentFolderAprendiz.objects.filter(aprendiz_id=aprendiz_id).values(
        "id", "name", "parent_id", "tipo", "documento__id", "documento__nom", "documento__archi"
    )

    folder_map = {}

    for nodo in nodos:
        nodo_id = nodo["id"]
        parent_id = nodo["parent_id"]

        # Construcción del nodo base
        nodo_data = {
            "id": nodo_id,
            "name": nodo["name"],
            "parent_id": parent_id,
            "tipo": nodo["tipo"],
            "children": []
        }

        # Si es un documento, añadir los datos adicionales
        if nodo["tipo"] == "documento":
            nodo_data.update({
                "documento_id": nodo["documento__id"],
                "documento_nombre": nodo["documento__nom"],
                "url": nodo["documento__archi"],  # La URL del archivo
            })

        # Guardar en el mapa de nodos
        folder_map[nodo_id] = nodo_data

    root_nodes = []

    # Construcción de la jerarquía
    for nodo in folder_map.values():
        parent_id = nodo["parent_id"]
        if parent_id:
            if parent_id in folder_map:
                folder_map[parent_id]["children"].append(nodo)
        else:
            root_nodes.append(nodo)

    return JsonResponse(root_nodes, safe=False)


@login_required
def descargar_portafolio_aprendiz_zip(request, aprendiz_id):
    # Obtener todos los nodos del árbol
    nodos = T_DocumentFolderAprendiz.objects.filter(
        aprendiz_id=aprendiz_id).select_related('documento')

    folder_map = {}
    doc_map = {}

    for nodo in nodos:
        folder_map[nodo.id] = {
            "id": nodo.id,
            "name": nodo.name,
            "parent_id": nodo.parent_id,
            "tipo": nodo.tipo,
            "children": [],
        }
        if nodo.tipo == "documento" and nodo.documento:
            doc_map[nodo.id] = nodo.documento

    # Construir jerarquía de carpetas
    root_nodes = []
    for nodo in folder_map.values():
        parent_id = nodo["parent_id"]
        if parent_id:
            folder_map[parent_id]["children"].append(nodo)
        else:
            root_nodes.append(nodo)

    # Función recursiva para añadir al ZIP
    def agregar_a_zip(zip_file, nodo, ruta):
        ruta_actual = os.path.join(ruta, nodo["name"])
        if nodo["tipo"] == "carpeta":
            for hijo in nodo["children"]:
                agregar_a_zip(zip_file, hijo, ruta_actual)
        elif nodo["tipo"] == "documento" and nodo["id"] in doc_map:
            doc = doc_map[nodo["id"]]
            if doc.archi and os.path.isfile(doc.archi.path):
                with open(doc.archi.path, 'rb') as f:
                    file_data = f.read()
                nombre_archivo = f"{doc.nom or os.path.basename(doc.archi.name)}"
                zip_file.writestr(os.path.join(
                    ruta, nombre_archivo), file_data)

    # Crear archivo ZIP en memoria
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for nodo in root_nodes:
            agregar_a_zip(zip_file, nodo, "")

    buffer.seek(0)

    AuditLog.objects.create(
        user=request.user,
        action="download",
        content_type=ContentType.objects.get_for_model(T_DocumentFolderAprendiz),
        object_id=None,
        related_id=aprendiz_id,
        related_type="aprendiz",
        extra_data=f"Descargó el portafolio completo del aprendiz {aprendiz_id} "
                   f"con {nodos.count()} nodos (documentos y carpetas)."
    )
    
    AuditLog.objects.create(
        user=request.user,
        action="download",
        content_type=ContentType.objects.get_for_model(T_DocumentFolderAprendiz),
        object_id=None,
        related_id=nodos[0].aprendiz.ficha_id,
        related_type="ficha",
        extra_data=f"Descargó el portafolio completo del aprendiz {aprendiz_id} "
                   f"con {nodos.count()} nodos (documentos y carpetas)."
    )

    response = HttpResponse(buffer, content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename=portafolio_aprendiz_{aprendiz_id}.zip'
    return response


@login_required
def descargar_portafolios_ficha_zip(request, ficha_id):
    # Obtener todos los aprendices de la ficha
    aprendices = T_apre.objects.filter(ficha_id=ficha_id)

    # Crear archivo ZIP en memoria
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for aprendiz in aprendices:
            nodos = T_DocumentFolderAprendiz.objects.filter(
                aprendiz=aprendiz).select_related('documento')

            folder_map = {}
            doc_map = {}

            for nodo in nodos:
                folder_map[nodo.id] = {
                    "id": nodo.id,
                    "name": nodo.name,
                    "parent_id": nodo.parent_id,
                    "tipo": nodo.tipo,
                    "children": [],
                }
                if nodo.tipo == "documento" and nodo.documento:
                    doc_map[nodo.id] = nodo.documento

            # Construir jerarquía
            root_nodes = []
            for nodo in folder_map.values():
                if nodo["parent_id"]:
                    folder_map[nodo["parent_id"]]["children"].append(nodo)
                else:
                    root_nodes.append(nodo)

            # Función recursiva para añadir al ZIP
            def agregar_a_zip(zip_file, nodo, ruta):
                if nodo["tipo"] == "carpeta":
                    ruta_actual = os.path.join(ruta, nodo["name"])
                    for hijo in nodo["children"]:
                        agregar_a_zip(zip_file, hijo, ruta_actual)
                elif nodo["tipo"] == "documento" and nodo["id"] in doc_map:
                    doc = doc_map[nodo["id"]]
                    if doc.archi and os.path.isfile(doc.archi.path):
                        with open(doc.archi.path, 'rb') as f:
                            file_data = f.read()
                        nombre_archivo = f"{doc.nom or os.path.basename(doc.archi.name)}"
                        # Guardar directamente en la ruta del aprendiz, sin subcarpeta adicional
                        zip_file.writestr(os.path.join(
                            ruta, nombre_archivo), file_data)

            # Carpeta por aprendiz
            carpeta_aprendiz = f"{aprendiz.perfil.dni}_{aprendiz.perfil.nom}_{aprendiz.perfil.apelli}".replace(
                ' ', '_')
            for nodo in root_nodes:
                agregar_a_zip(zip_file, nodo, carpeta_aprendiz)

    buffer.seek(0)
    
    AuditLog.objects.create(
        user=request.user,
        action="download",
        content_type=ContentType.objects.get_for_model(T_DocumentFolderAprendiz),
        object_id=None,
        related_id=ficha_id,
        related_type="ficha",
        extra_data=f"Descargó el portafolio completo de todos los aprendices para la ficha {ficha_id} "
                   f"con {nodos.count()} nodos (documentos y carpetas)."
    )

    response = HttpResponse(buffer, content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename=portafolios_ficha_{ficha_id}.zip'
    return response


@login_required
def eliminar_documento_portafolio_aprendiz(request, documento_id):
    if request.method != "DELETE":
        return JsonResponse({"status": "error", "error": "Método no permitido"}, status=405)

    documento = get_object_or_404(T_DocumentFolderAprendiz, id=documento_id)

    nombre_doc = documento.name if (documento.name) else "Documento sin archivo"
    extra_data = (
        f"Se elimino el documento '{nombre_doc}' "
        f"en la carpeta {documento.parent_id} "
    )

    if documento.documento and documento.documento.archi:
        archivo_a_eliminar = documento.documento.archi.name
        if archivo_a_eliminar:
            default_storage.delete(archivo_a_eliminar)
        documento.documento.delete()

    documento.delete()
    
    AuditLog.objects.create(
        user=request.user,
        content_type=ContentType.objects.get_for_model(documento),
        object_id=documento.id,
        action="delete",
        related_id = documento.aprendiz_id,
        related_type = "aprendiz",
        extra_data=extra_data
    )


    return JsonResponse({"status": "success", "message": "Eliminado correctamente"}, status=200)


@login_required
def obtener_hijos_carpeta_aprendiz(request, carpeta_id):
    try:
        nodos = T_DocumentFolderAprendiz.objects.filter(parent_id=carpeta_id).values(
            "id", "name", "parent_id", "tipo", "documento__id", "documento__nom", "documento__archi"
        )

        children = []

        for nodo in nodos:
            data = {
                "id": nodo["id"],
                "name": nodo["name"],
                "parent_id": nodo["parent_id"],
                "tipo": nodo["tipo"],
                "children": []
            }

            if nodo["tipo"] == "documento":
                data.update({
                    "documento_id": nodo["documento__id"],
                    "documento_nombre": nodo["documento__nom"],
                    "url": nodo["documento__archi"],
                })

            children.append(data)

        return JsonResponse(children, safe=False)

    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


@login_required
@bloquear_si_consulta
def crear_encuentro(request, ficha_id):
    ficha = get_object_or_404(T_ficha, id=ficha_id)
    if request.method == 'POST':
        encuentro_form = EncuentroForm(request.POST)
        encuapre_form = EncuApreForm(request.POST, ficha=ficha)
        if encuentro_form.is_valid() and encuapre_form.is_valid():

            new_encuentro = encuentro_form.save(commit=False)

            fase_ficha = T_fase_ficha.objects.filter(
                ficha=ficha, vige='1').first()

            new_encuentro.fase = fase_ficha
            new_encuentro.ficha = ficha
            new_encuentro.save()
            aprendices_seleccionados = encuapre_form.cleaned_data['aprendices']
            aprendices_ficha = T_apre.objects.filter(ficha=ficha)
            aprendices_seleccionados_set = set(aprendices_seleccionados)

            for aprendiz in aprendices_ficha:
                if aprendiz in aprendices_seleccionados_set:
                    T_encu_apre.objects.update_or_create(
                        encu=new_encuentro,
                        apre=aprendiz,
                        defaults={'prese': 'No'}
                    )
                else:
                    T_encu_apre.objects.update_or_create(
                        encu=new_encuentro,
                        apre=aprendiz,
                        defaults={'prese': 'Si'}
                    )

            for aprendiz_encu in aprendices_seleccionados:
                T_encu_apre.objects.update_or_create(
                    encu=new_encuentro,
                    apre=aprendiz_encu,
                    defaults={'prese': 'No'}
                )

            return JsonResponse({'status': 'success', 'message': 'Encuentro creado con exito'}, status=200)
        else:
            errores_custom = []

            for field, errors_list in encuentro_form.errors.get_json_data().items():
                nombre_campo = encuentro_form.fields[field].label or field.capitalize(
                )
                for err in errors_list:
                    errores_custom.append(
                        f"<strong>{nombre_campo}</strong>: {err['message']}")

            for field, errors_list in encuapre_form.errors.get_json_data().items():
                nombre_campo = encuapre_form.fields[field].labels or field.capitalize(
                )
                for err in errors_list:
                    errores_custom.append(
                        f"<strong>{nombre_campo}</strong>: {err['message']}")

            return JsonResponse({'status': 'error', 'message': 'Errores en el formulario', 'errors': '<br>'.join(errores_custom)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Metodo no permitido'}, status=405)


@login_required
def obtener_encuentros(request, ficha_id):
    encuentros = T_encu.objects.filter(ficha_id=ficha_id)
    data = [
        {
            'id': encu.id,
            'fecha': encu.fecha,
            'tema': encu.tema,
            'lugar': encu.lugar
        } for encu in encuentros
    ]
    return JsonResponse(data, safe=False)


@require_GET
@login_required
def obtener_encuentro(request, encuentro_id):
    encuentro = T_encu.objects.get(pk=encuentro_id)

    if not encuentro:
        return JsonResponse({'status': 'error', 'message': 'Encuentro no encontrado'}, status=404)

    ausentes_qs = T_encu_apre.objects.filter(encu=encuentro, prese="No")
    ausentes_ids = list(ausentes_qs.values_list('apre_id', flat=True))
    data = {
        'tema': encuentro.tema,
        'lugar': encuentro.lugar,
        'fecha': encuentro.fecha.strftime('%Y-%m-%d'),
        'ausentes': ausentes_ids
    }
    return JsonResponse(data, safe=False)


@require_POST
@login_required
@bloquear_si_consulta
def editar_encuentro(request, encuentro_id):
    try:
        encuentro = T_encu.objects.get(pk=encuentro_id)
    except T_encu.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Encuentro no encontrado'}, status=404)

    tema = request.POST.get('tema')
    lugar = request.POST.get('lugar')
    fecha_str = request.POST.get('fecha')

    try:
        fecha_naive = datetime.strptime(fecha_str, "%Y-%m-%d")
        fecha_aware = timezone.make_aware(fecha_naive)
    except (ValueError, TypeError):
        return JsonResponse({'status': 'error', 'message': 'Formato de fecha inválido'}, status=400)

    # Asignar todos los campos antes de guardar
    encuentro.tema = tema
    encuentro.lugar = lugar
    encuentro.fecha = fecha_aware
    encuentro.save()

    ausentes = request.POST.getlist('aprendices')

    registros = T_encu_apre.objects.filter(encu=encuentro)
    for reg in registros:
        reg.prese = "No" if str(reg.apre.id) in ausentes else "Si"
        reg.save()

    return JsonResponse({'status': 'success', 'message': 'Encuentro actualizado correctamente'}, status=200)


###############################################################################################################
#        VISTAS FICHA MASIVO
###############################################################################################################

# Función para generar contraseña aleatoria


def generar_contraseña(length=8):
    caracteres = string.ascii_letters + string.digits
    return ''.join(random.choice(caracteres) for _ in range(length))


@login_required
def cargar_fichas_masivo(request):
    return render(request, 'fichas_masivo_crear.html')


def formatear_error_csv(fila, errores_campos):
    fila_str = '\n  '.join(
        [f"{k}: '{v}'" for k, v in fila.items() if k != '__line__'])
    linea = fila.get('__line__', '?')
    return (
        f"⚠️ Error en la línea {linea} del archivo CSV:\n"
        f"----------------------------------------\n"
        f"Datos de la fila:\n  {fila_str}\n\n"
        f"Errores encontrados:\n" +
        '\n'.join(f"- {e}" for e in errores_campos) + "\n"
        f"----------------------------------------"
    )


def validar_campos_csv(fila, emails_csv=None, dnis_csv=None):
    errores = []

    emails_csv = emails_csv if emails_csv is not None else set()
    dnis_csv = dnis_csv if dnis_csv is not None else set()

    TIPOS_DNI_VALIDOS = {'ti', 'cc', 'pp', 'ce', 'ppt'}
    GENEROS_VALIDOS = {'H', 'M'}
    PARENTEZCOS_VALIDOS = {
        'padre', 'madre', 'abuelo', 'abuela', 'hermano', 'hermana', 'tio', 'tia', 'otro'
    }

    campos_requeridos = [
        'email', 'nom', 'apelli', 'tipo_dni', 'dni', 'tele',
        'dire', 'gene', 'fecha_naci', 'nom_repre', 'dni_repre',
        'tele_repre', 'dire_repre', 'mail_repre', 'parentezco'
    ]

    for campo in campos_requeridos:
        if campo not in fila or not fila[campo].strip():
            errores.append(f"Campo requerido faltante: '{campo}'")

    # -----------------------
    # Validaciones del aprendiz
    # -----------------------

    # Email aprendiz
    email = fila['email'].strip().lower()
    try:
        validate_email(email)
    except ValidationError:
        errores.append(f"Email del aprendiz inválido: '{email}'")
    else:
        if email in emails_csv:
            errores.append(
                f"Email del aprendiz duplicado en el archivo: '{email}'")
        elif T_perfil.objects.filter(mail=email).exists():
            errores.append(
                f"Email del aprendiz ya registrado en el sistema: '{email}'")
        else:
            emails_csv.add(email)

    # Nombres
    nom = fila['nom'].strip()
    if not (2 <= len(nom) <= 50):
        errores.append(
            f"Nombres del aprendiz inválidos (mínimo 2, máximo 50 caracteres): '{nom}'")

    # Apellidos
    apelli = fila['apelli'].strip()
    if not (2 <= len(apelli) <= 50):
        errores.append(
            f"Apellidos del aprendiz inválidos (mínimo 2, máximo 50 caracteres): '{apelli}'")

    # Tipo de documento
    tipo_dni = fila['tipo_dni'].strip().lower()
    if tipo_dni not in TIPOS_DNI_VALIDOS:
        errores.append(
            f"Tipo de documento inválido: '{tipo_dni}'. Debe ser uno de: {', '.join(sorted(TIPOS_DNI_VALIDOS))}")

    # DNI aprendiz
    dni = fila['dni'].strip()
    if not dni.isdigit() or int(dni) <= 0:
        errores.append(
            f"DNI del aprendiz inválido, debe ser un número: '{dni}'")
    elif not (6 <= len(dni) <= 15):
        errores.append(
            f"DNI del aprendiz inválido, longitud fuera de rango (6-15): '{dni}'")
    else:
        if dni in dnis_csv:
            errores.append(
                f"DNI del aprendiz duplicado en el archivo: '{dni}'")
        elif T_perfil.objects.filter(dni=dni).exists():
            errores.append(
                f"DNI del aprendiz ya registrado en el sistema: '{dni}'")
        else:
            dnis_csv.add(dni)

    # Teléfono aprendiz
    tele = fila['tele'].strip()
    if not tele.isdigit() or int(tele) <= 0:
        errores.append(
            f"Teléfono del aprendiz inválido, debe contener solo números: '{tele}'")
    elif not (7 <= len(tele) <= 15):
        errores.append(
            f"Teléfono del aprendiz inválido, longitud fuera de rango (7-15): '{tele}'")

    # Dirección
    dire = fila['dire'].strip()
    if len(dire) < 5:
        errores.append(
            f"Dirección del aprendiz demasiado corta (mínimo 5 caracteres): '{dire}'")

    # Género
    gene = fila['gene'].strip().upper()
    if gene not in GENEROS_VALIDOS:
        errores.append(f"Género inválido: '{gene}'. Debe ser 'H' o 'M'.")

    # Fecha de nacimiento
    try:
        datetime.strptime(fila['fecha_naci'].strip(), '%d/%m/%Y').date()
    except ValueError:
        errores.append(
            f"Fecha de nacimiento inválida: '{fila['fecha_naci']}'. Formato correcto: DD/MM/AAAA.")

    # -----------------------
    # Validaciones del representante
    # -----------------------

    # Nombre representante
    nom_repre = fila['nom_repre'].strip()
    if not (2 <= len(nom_repre) <= 50):
        errores.append(
            f"Nombre del representante inválido (mínimo 2, máximo 50 caracteres): '{nom_repre}'")

    # DNI representante
    dni_repre = fila['dni_repre'].strip()
    if not dni_repre.isdigit() or int(dni_repre) <= 0:
        errores.append(
            f"DNI del representante inválido, debe contener solo números positivos: '{dni_repre}'")
    elif not (6 <= len(dni_repre) <= 15):
        errores.append(
            f"DNI del representante inválido, longitud fuera de rango (6-15): '{dni_repre}'")

    # Teléfono representante
    tele_repre = fila['tele_repre'].strip()
    if not tele_repre.isdigit() or int(tele_repre) <= 0:
        errores.append(
            f"Teléfono del representante inválido, debe contener solo números positivos: '{tele_repre}'")
    elif not (7 <= len(tele_repre) <= 15):
        errores.append(
            f"Teléfono del representante inválido, longitud fuera de rango (7-15): '{tele_repre}'")

    # Dirección representante
    dire_repre = fila['dire_repre'].strip()
    if len(dire_repre) < 5:
        errores.append(
            f"Dirección del representante demasiado corta (mínimo 5 caracteres): '{dire_repre}'")

    # Email representante
    email_repre = fila['mail_repre'].strip()
    try:
        validate_email(email_repre)
    except ValidationError:
        errores.append(
            f"Correo electrónico del representante inválido: '{email_repre}'")

    # Parentezco
    paren = fila['parentezco'].strip().lower()
    if paren not in PARENTEZCOS_VALIDOS:
        errores.append(
            f"Parentesco inválido: '{paren}'. Debe ser uno de: {', '.join(sorted(PARENTEZCOS_VALIDOS))}")

    return errores


@login_required
@require_POST
def cargar_fichas(request):
    errores = []
    resumen = {
        "insertados": 0,
        "errores": 0,
        "duplicados_dni": []
    }

    archivo = request.FILES.get('archivo')
    if not archivo:
        return JsonResponse({
            "success": False,
            "message": "No se recibió ningún archivo.",
            "errores": ["No se recibió ningún archivo."]
        }, status=400)

    if not archivo.name.lower().endswith('.csv'):
        return JsonResponse({
            "success": False,
            "message": "Solo se permiten archivos CSV (.csv)",
            "errores": ["Solo se permiten archivos CSV (.csv)"]
        }, status=400)

    allowed_mime_types = ['text/csv', 'application/csv', 'text/plain']
    if archivo.content_type not in allowed_mime_types:
        return JsonResponse({
            "success": False,
            "message": "Tipo de archivo no válido (solo CSV)",
            "errores": ["Tipo de archivo no válido (solo CSV)"]
        }, status=400)

    datos_csv = TextIOWrapper(archivo.file, encoding='utf-8-sig')
    lector_csv = csv.DictReader(datos_csv, delimiter=';')

    lector = [
        fila for fila in lector_csv
        if any(campo.strip() for campo in fila.values())
    ]

    if len(lector) > 60:
        return JsonResponse({
            "success": False,
            "message": "El archivo contiene más de 60 registros.",
            "errores": ["El archivo no debe exceder los 60 registros."]
        }, status=400)

    try:
        with transaction.atomic():
            aprendices_creados = []
            emails_csv = set()
            dnis_csv = set()

            for i, fila in enumerate(lector, start=2):

                fila['__line__'] = i

                try:

                    errores_fila = validar_campos_csv(
                        fila, emails_csv, dnis_csv)

                    if errores_fila:
                        error_msg = formatear_error_csv(fila, errores_fila)
                        raise ValidationError(error_msg)

                    dni = fila['dni']
                    fecha_naci = datetime.strptime(
                        fila['fecha_naci'].strip(), '%d/%m/%Y').date()
                    base_username = (fila['nom'][:3] +
                                     fila['apelli'][:3]).lower()
                    username = base_username
                    i = 1
                    while User.objects.filter(username=username).exists():
                        username = f"{base_username}{i}"
                        i += 1

                    # contraseña = generar_contraseña()
                    user = User.objects.create_user(
                        username=username,
                        password=str(dni),
                        email=fila['email']
                    )

                    perfil = T_perfil.objects.create(
                        user=user,
                        nom=fila['nom'],
                        apelli=fila['apelli'],
                        tipo_dni=fila['tipo_dni'],
                        dni=dni,
                        tele=fila['tele'],
                        dire=fila['dire'],
                        gene=fila['gene'],
                        mail=fila['email'],
                        fecha_naci=fecha_naci,
                        rol="aprendiz"
                    )
                    try:
                        perfil.full_clean()
                    except ValidationError as e:
                        errores = []
                        for campo, mensajes in e.message_dict.items():
                            valor = getattr(perfil, campo, '')
                            for msg in mensajes:
                                errores.append(
                                    f"{msg} Valor recibido en campo '{campo}': '{valor}'")
                        raise ValidationError(
                            formatear_error_csv(fila, errores))

                    validate_email(fila['mail_repre'].strip())

                    representante_legal = T_repre_legal.objects.filter(
                        dni=fila['dni_repre']
                    ).first()

                    if not representante_legal:
                        representante_legal = T_repre_legal(
                            nom=fila['nom_repre'],
                            dni=fila['dni_repre'],
                            tele=fila['tele_repre'],
                            dire=fila['dire_repre'],
                            mail=fila['mail_repre'],
                            paren=fila['parentezco']
                        )
                        try:
                            representante_legal.full_clean()
                        except ValidationError as e:
                            errores = []
                            for campo, mensajes in e.message_dict.items():
                                valor = getattr(representante_legal, campo, '')
                                for msg in mensajes:
                                    errores.append(
                                        f"{msg} Valor recibido en campo '{campo}': '{valor}'")
                            raise ValidationError(
                                formatear_error_csv(fila, errores))
                        representante_legal.save()

                    aprendiz = T_apre.objects.create(
                        cod="z",
                        esta="activo",
                        perfil=perfil,
                        repre_legal=representante_legal,
                        usu_crea=request.user
                    )
                    aprendiz.full_clean()
                    aprendices_creados.append(aprendiz)
                    resumen["insertados"] += 1

                except ValidationError as ve:
                    raise ve
                except Exception as fila_error:
                    raise ValidationError(
                        f"Error inesperado en la fila {fila.get('__line__', '?')}: {str(fila_error)}")

            # Crear ficha y grupo
            cod_ficha = request.POST.get('num_ficha', '').strip()
            fase = request.POST.get('fase_actual')
            colegio = request.POST.get('colegio')
            centro_forma = request.POST.get('centro_forma')
            programa = request.POST.get('programa')

            if not fase or not colegio or not centro_forma or not programa:
                raise ValidationError("Faltan datos para crear la ficha.")

            try:
                programaf = T_progra.objects.get(id=programa)
            except T_progra.DoesNotExist:
                raise ValidationError("Programa de formación no encontrado.")

            if cod_ficha and T_ficha.objects.filter(num=cod_ficha).exists():
                raise ValidationError("Numero de ficha ya existe.")

            centro = T_centro_forma.objects.filter(pk=centro_forma).first()
            institucion = T_insti_edu.objects.filter(pk=colegio).first()
            grupo = T_grupo.objects.create(
                esta='Masivo',
                fecha_crea=timezone.now(),
                autor=request.user,
                num_apre_poten=len(aprendices_creados),
                centro=centro,
                insti=institucion,
                progra=programaf
            )
            grupo.full_clean()
            grupo.save()

            departamento = institucion.muni.nom_departa if institucion and institucion.muni else None

            gestor_depa = T_gestor_depa.objects.filter(
                depa=departamento).select_related('gestor').first()

            if not gestor_depa:
                gestor = T_gestor.objects.get(pk=1)
            else:
                gestor = gestor_depa.gestor

            T_gestor_grupo.objects.create(
                fecha_crea=timezone.now(),
                autor=request.user,
                gestor=gestor,
                grupo=grupo
            )

            documentos_matricula = [
                'Documento de Identidad del aprendiz',
                'Registro civil',
                'Certificado de Afiliación de salud',
                'Formato de Tratamiento de Datos del Menor de Edad',
                'Compromiso del Aprendiz',
            ]

            perfili = T_perfil.objects.get(user=request.user)
            instructor = T_instru.objects.get(perfil=perfili)

            ficha = T_ficha.objects.create(
                num=cod_ficha,
                grupo=grupo,
                fecha_aper=timezone.now(),
                fecha_cierre=None,
                insti=institucion,
                centro=centro,
                progra=programaf,
                num_apre_proce=len(aprendices_creados),
                num_apre_forma=0,
                num_apre_pendi_regi=len(aprendices_creados),
                esta="Activo",
                instru=instructor
            )

            for aprendiz in aprendices_creados:
                aprendiz.grupo = grupo
                aprendiz.ficha = ficha
                aprendiz.esta_docu = "Pendiente"
                aprendiz.save()

                for documento in documentos_matricula:
                    T_prematri_docu.objects.create(
                        nom=documento,
                        apren=aprendiz,
                        esta="Pendiente",
                        vali="0"
                    )

            for f in range(1, int(fase)):
                T_fase_ficha.objects.create(
                    fase_id=f,
                    ficha=ficha,
                    fecha_ini=timezone.now(),
                    instru=instructor,
                    vige=0
                )

            T_fase_ficha.objects.create(
                fase_id=fase,
                ficha=ficha,
                fecha_ini=timezone.now(),
                instru=instructor,
                vige=1
            )

            crear_datos_prueba(ficha.id)

            for aprendiz in aprendices_creados:
                crear_datos_prueba_aprendiz(aprendiz.id)

    except ValidationError as e:
        return JsonResponse({
            "success": False,
            "message": "Se detectaron errores, operación revertida.",
            "errores": e.messages,
            "resumen": resumen
        }, status=400)

    return JsonResponse({
        "success": True,
        "message": (
            f"✅ Ficha y aprendices creados correctamente.\n\n"
            f"Ficha creada: ID: {ficha.id}, Numero: {ficha.num or 'Sin numero'}\n"
            f"Grupo creado:\n  ID: {grupo.id}"
            f"\nAprendices insertados: {resumen['insertados']}\n"
            f"{'DNIs duplicados encontrados: ' + ', '.join(resumen['duplicados_dni']) if resumen['duplicados_dni'] else ''}"
        ),
        "resumen": resumen,
        "errores": errores
    }, status=201)


@require_GET
@login_required
def obtener_opciones_fichas_masivo_departamentos(request):
    departamentos = T_departa.objects.all().distinct().values('id', 'nom_departa')
    data = [{'id': d['id'], 'nom': d['nom_departa']} for d in departamentos]
    return JsonResponse(list(data), safe=False)


@require_GET
@login_required
def obtener_opciones_fichas_masivo_municipios(request):
    departamento_id = request.GET.get('departamento')

    queryset = T_munici.objects.all()

    if departamento_id:
        queryset = queryset.filter(nom_departa_id=departamento_id)

    municipios = queryset.distinct().values('id', 'nom_munici')
    data = [{'id': m['id'], 'nom': m['nom_munici']} for m in municipios]
    return JsonResponse(list(data), safe=False)


@require_GET
@login_required
def obtener_opciones_fichas_masivo_colegios(request):
    municipio_id = request.GET.get('municipio')

    queryset = T_insti_edu.objects.all()

    if municipio_id:
        queryset = queryset.filter(muni_id=municipio_id)
    else:
        queryset = T_insti_edu.objects.none()

    instituciones = queryset.distinct().values('id', 'nom')
    data = [{'id': i['id'], 'nom': i['nom']} for i in instituciones]
    return JsonResponse(list(data), safe=False)


@require_GET
@login_required
def obtener_opciones_fichas_masivo_centros(request):
    centros = T_centro_forma.objects.all().distinct().values('id', 'nom')
    data = [{'id': c['id'], 'nom': c['nom']} for c in centros]
    return JsonResponse(list(data), safe=False)


@require_GET
@login_required
def obtener_opciones_fichas_masivo_programas(request):
    programas = T_progra.objects.all().distinct().values('id', 'nom')
    data = [{'id': p['id'], 'nom': p['nom']} for p in programas]
    return JsonResponse(list(data), safe=False)

###############################################################################################################
#        VISTAS APRENDIZ
###############################################################################################################


@login_required
def panel_aprendiz(request):
    # Obtener el perfil del usuario logueado
    perfil = T_perfil.objects.filter(user=request.user).first()

    # Verificar que el perfil existe
    if perfil:
        # Buscar el aprendiz asociado con ese perfil
        aprendiz = T_apre.objects.filter(perfil=perfil).first()

        # Verificar si el aprendiz tiene ficha asignada
        ficha = aprendiz.ficha if aprendiz else None

        # Obtener el listado de documentos del aprendiz
        documentos = T_prematri_docu.objects.filter(
            apren=aprendiz) if aprendiz else []

        total_documentos = 0
        for documento in documentos:
            if documento.esta == "Cargado":
                total_documentos += 1
    else:
        ficha = None
        documentos = []
        total_documentos = 0
    return render(request, 'panel_aprendiz.html', {'ficha': ficha, 'documentos': documentos, 'total_documentos': total_documentos})

###############################################################################################################
#        VISTAS PROGRAMA
###############################################################################################################


@login_required
def listar_programas(request):
    programas = T_progra.objects.all()
    return render(request, 'programas.html', {'programas': programas})


@login_required
@bloquear_si_consulta
def crear_programa(request):
    if request.method == 'POST':
        programa_form = ProgramaForm(request.POST)
        if programa_form.is_valid():
            programa_form.save()
            return redirect('programas')
    else:
        programa_form = ProgramaForm()
    return render(request, 'crear_programa.html', {'programa_form': programa_form})

###############################################################################################################
#        VISTAS COMPETENCIA
###############################################################################################################


@login_required
def competencias(request):
    return render(request, 'competencias.html')


###############################################################################################################
#        VISTAS RAPS
###############################################################################################################

@login_required
def listar_raps(request):
    return render(request, 'raps.html')

@login_required
def obtener_opciones_fases_raps(request):
    fases = T_fase.objects.filter(
        t_raps_fase__isnull=False
    ).distinct().values_list('nom', flat=True)

    return JsonResponse(list(fases), safe=False)


@login_required
def obtener_opciones_programas_raps(request):
    programas = T_progra.objects.all().distinct().values_list('nom', flat=True)
    return JsonResponse(list(programas), safe=False)


@login_required
def obtener_opciones_competencias_raps(request):
    competencias = T_compe.objects.filter(
        t_raps__isnull=False
    ).distinct().values_list('nom', flat=True)

    return JsonResponse(list(competencias), safe=False)


@login_required
def obtener_rap(request, rap_id):
    rap = T_raps.objects.filter(pk=rap_id).first()
    if rap:
        fases = list(rap.fase.values('id', 'nom'))
        data = {
            'id': rap.id,
            'nom': rap.nom,
            'compe': rap.compe.id,
            'fases': fases
        }
        return JsonResponse(data)
    return JsonResponse({'status': 'error', 'message': 'RAP no encontrado'}, status=404)


@login_required
def obtener_opciones_competencias(request):
    competencias = T_compe.objects.all().values('id', 'nom')
    return JsonResponse(list(competencias), safe=False)


@login_required
@bloquear_si_consulta
def editar_rap(request, rap_id):
    rap = T_raps.objects.filter(pk=rap_id).first()

    if not rap:
        return JsonResponse({'status': 'error', 'message': 'RAP no encontrado'}, status=404)

    if request.method == 'POST':
        form_rap = RapsForm(request.POST, instance=rap)

        if form_rap.is_valid():
            nom = form_rap.cleaned_data['nom']
            if T_raps.objects.filter(nom=nom).exclude(pk=rap_id).exists():
                return JsonResponse({'status': 'error', 'message': 'Ya existe un RAP con el nombre indicado'}, status=400)

            form_rap.save
            return JsonResponse({'status': 'success', 'message': 'RAP actualizado con éxito'}, status=200)
        else:
            return JsonResponse({'status': 'error', 'message': 'Error al actualizar el RAP', 'errors': form_rap.errors}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)


@login_required
def eliminar_doc(request, documento_id):
    if request.method == "DELETE":
        try:
            documento = T_DocumentFolder.objects.get(id=documento_id)

            # Eliminar archivo físico antes de borrar el modelo relacionado
            if documento.documento and documento.documento.archi:
                archivo_a_eliminar = documento.documento.archi.name
                if archivo_a_eliminar:
                    default_storage.delete(archivo_a_eliminar)
                documento.documento.delete()

            documento.delete()

            return JsonResponse({"success": True, "message": "Documento eliminado exitosamente."}, status=200)

        except T_DocumentFolder.DoesNotExist:
            return JsonResponse({"success": False, "message": "Documento no encontrado."}, status=404)

    return JsonResponse({"success": False, "message": "Método no permitido."}, status=405)


@login_required
def listar_estudiantes(request, ficha_id):
    estudiantes = T_apre.objects.filter(ficha_id=ficha_id)
    data = [
        {
            "id": estudiante.id,
            "nombre": estudiante.perfil.nom,
            "apellido": estudiante.perfil.apelli,
        }
        for estudiante in estudiantes
    ]
    return JsonResponse(data, safe=False)

# Vista para cargar los municipios según el departamento


@login_required
def get_municipios(request, departamento_id):
    municipio_qs = T_munici.objects.filter(nom_departa_id=departamento_id)
    municipios = list(municipio_qs.values('id', 'nom_munici'))
    return JsonResponse(municipios, safe=False)

# Vista para cargar las instituciones según el municipio


@login_required
def get_instituciones(request, municipio_id):
    instituciones_qs = T_insti_edu.objects.filter(muni_id=municipio_id)
    instituciones = list(instituciones_qs.values('id', 'nom'))
    return JsonResponse(instituciones, safe=False)

# Vista para cargar los municipios según el departamento

@login_required
def get_centros(request, departamento_id):
    centro_qs = T_centro_forma.objects.filter(depa_id=departamento_id)
    centros = list(centro_qs.values('id', 'nom'))
    return JsonResponse(centros, safe=False)

@login_required
def detalle_programa(request, programa_id):
    programa = T_progra.objects.get(pk=programa_id)

    competencias = []
    for compe in T_compe.objects.filter(progra=programa):
        fases = T_fase.objects.filter(
            t_compe_fase__compe=compe).values_list('nom', flat=True)
        competencias.append({
            'nom': compe.nom,
            'fase': list(fases),
        })

    raps = []
    for compe in T_compe.objects.filter(progra=programa):
        for rap in compe.t_raps_set.all():
            raps.append({
                'nom': rap.nom,
                'compe': compe.nom,
            })

    return JsonResponse({
        'cod_prog': programa.cod_prog,
        'nom': programa.nom,
        'nomd': programa.nomd,
        'competencias': competencias,
        'raps': raps,
    })

def link_callback(uri, rel):
    result = finders.find(uri.replace(settings.STATIC_URL, ""))
    if result:
        return result
    raise Exception(f"Media URI must start with {settings.STATIC_URL}")

@login_required
def generar_acta_asistencia(request):
    ficha_id = request.GET.get('ficha_id')
    formato = request.GET.get('formato')

    ficha = T_ficha.objects.get(id=ficha_id)
    encuentros = T_encu.objects.filter(ficha=ficha).order_by('fecha')
    aprendices = T_apre.objects.filter(ficha=ficha)

    asistencias = {
        apre.id: {
            encu.id: "No"
            for encu in encuentros
        } for apre in aprendices
    }

    registros = T_encu_apre.objects.filter(
        encu__in=encuentros, apre__in=aprendices)

    for reg in registros:
        asistencias[reg.apre.id][reg.encu.id] = reg.prese

    if formato == 'pdf':
        template = get_template("reportes/acta_asistencia_general.html")
        tabla = []
        for apre in aprendices:
            fila = {
                'apre': apre,
                'asistencias': []
            }
            for encu in encuentros:
                prese = asistencias.get(apre.id, {}).get(encu.id, "No")
                fila['asistencias'].append("Sí" if prese == "Si" else "No")
            tabla.append(fila)
        html = template.render({
            'logo_url': settings.STATIC_URL + 'images/imagenhome.png',
            'tabla': tabla,
            'ficha': ficha,
            'encuentros': encuentros,
            'aprendices': aprendices,
            'asistencias': asistencias,
        })
        buffer = BytesIO()
        pisa.CreatePDF(html, dest=buffer, link_callback=link_callback)
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename=f"Acta_Asistencia_{ficha.id}.pdf")

    elif formato == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Asistencia"

        # Estilos
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        fill_header = PatternFill(
            start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        fill_asistencias_col = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_fallas_col = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fill_fila_asistentes = PatternFill(
            start_color="A9D08E", end_color="A9D08E", fill_type="solid")
        fill_fila_fallas = PatternFill(
            start_color="F4B084", end_color="F4B084", fill_type="solid")

        if ficha.instru and ficha.instru.perfil:
            nombre_instructor = f"{ficha.instru.perfil.nom.upper()} {ficha.instru.perfil.apelli.upper()}"
        else:
            nombre_instructor = "INSTRUCTOR NO ASIGNADO"

        # Info general
        ws['A1'] = "1 = Asistencia (Sí)"
        ws['C1'] = f"Ficha: {ficha.num}"
        ws['A2'] = "0 = No asistencia (No)"
        ws['C2'] = f"Programa: {ficha.progra.nom}"
        ws['C3'] = f"Instructor: {nombre_instructor}"
        ws['A5'] = "PLANILLA  DE ASISTENCIA"

        headers = ["No.", "TIPO ID", "NUMERO ID",
                   "APELLIDOS Y NOMBRES", "E-MAIL", "TELEFONO"]
        fechas = list(T_encu.objects.filter(ficha=ficha).order_by(
            'fecha').values_list('id', 'fecha'))
        headers += [fecha.strftime("%d-%b-%y") for _, fecha in fechas]
        headers += ["Asistencias", "Fallas"]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = fill_header
            cell.border = thin_border

        conteo_por_fecha = {encu_id: {'asistencias': 0, 'faltas': 0}
                            for encu_id, _ in fechas}
        aprendices = aprendices.order_by('perfil__apelli', 'perfil__nom')

        for i, apre in enumerate(aprendices, start=1):
            row_num = i + 6
            perfil = apre.perfil
            asistencias = T_encu_apre.objects.filter(
                apre=apre, encu__ficha=ficha).select_related('encu')
            asistencia_por_id_encu = {
                a.encu_id: 1 if a.prese.strip().lower() == 'si' else 0 for a in asistencias
            }

            data = [
                i,
                perfil.tipo_dni,
                perfil.dni,
                f"{perfil.apelli} {perfil.nom}",
                perfil.mail,
                perfil.tele
            ]
            for col_idx, val in enumerate(data, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.border = thin_border

            total_asistencias = 0
            for j, (encu_id, _) in enumerate(fechas, start=1):
                valor = asistencia_por_id_encu.get(encu_id, 0)
                col_idx = 6 + j
                cell = ws.cell(row=row_num, column=col_idx, value=valor)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                total_asistencias += valor

                if valor == 1:
                    conteo_por_fecha[encu_id]['asistencias'] += 1
                else:
                    conteo_por_fecha[encu_id]['faltas'] += 1

            total_fallas = len(fechas) - total_asistencias
            asist_cell = ws.cell(row=row_num, column=6 +
                                 len(fechas) + 1, value=total_asistencias)
            fallas_cell = ws.cell(row=row_num, column=6 +
                                  len(fechas) + 2, value=total_fallas)

            asist_cell.fill = fill_asistencias_col
            fallas_cell.fill = fill_fallas_col
            asist_cell.border = thin_border
            fallas_cell.border = thin_border
            asist_cell.alignment = fallas_cell.alignment = Alignment(
                horizontal='center')

        fila_totales_asistentes = len(aprendices) + 7
        fila_totales_fallas = len(aprendices) + 8

        ws.cell(row=fila_totales_asistentes, column=6).value = "Asistentes"
        ws.cell(row=fila_totales_fallas, column=6).value = "No asistentes"

        for j, (encu_id, _) in enumerate(fechas, start=1):
            asistentes = conteo_por_fecha[encu_id]['asistencias']
            no_asistentes = conteo_por_fecha[encu_id]['faltas']

            for fila, valor, fill in [
                (fila_totales_asistentes, asistentes, fill_fila_asistentes),
                (fila_totales_fallas, no_asistentes, fill_fila_fallas),
            ]:
                cell = ws.cell(row=fila, column=6 + j, value=valor)
                cell.font = Font(bold=True)
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            # o set un ancho fijo si no funciona bien
            ws.column_dimensions[get_column_letter(col)].auto_size = True

        # Guardar y retornar
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename=f"Planilla_Asistencia_{ficha.num}.xlsx")

    return HttpResponse("Formato no válido", status=400)

@login_required
def generar_acta_asistencia_aprendiz(request):
    aprendiz_id = request.GET.get('aprendiz_id')
    aprendiz = T_apre.objects.get(id=aprendiz_id)
    asistencias = T_encu_apre.objects.filter(apre=aprendiz)

    context = {
        'aprendiz': aprendiz,
        'asistencias': asistencias
    }

    template = get_template("reportes/acta_asistencia_aprendiz.html")
    html = template.render({
        'aprendiz': aprendiz,
        'asistencias': asistencias,
    })

    buffer = BytesIO()
    pisa.CreatePDF(html, dest=buffer)
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename=f"Acta_Asistencia_{aprendiz.perfil.nom}_{aprendiz.perfil.apelli}.pdf")

@login_required
def detalle_encuentro(request, encuentro_id):
    encuentro = T_encu.objects.get(id=encuentro_id)

    total_aparticipantes = T_encu_apre.objects.filter(encu=encuentro).count()

    asistencias = T_encu_apre.objects.filter(encu=encuentro)

    aprendices_asistieron = []
    aprendices_faltaron = []

    for asistencia in asistencias:
        aprendiz = asistencia.apre
        aprendiz_data = {
            'id': aprendiz.id,
            'nombre': f"{aprendiz.perfil.nom} {aprendiz.perfil.apelli}"
        }
        if asistencia.prese == "Si":
            aprendices_asistieron.append(aprendiz_data)
        elif asistencia.prese == "No":
            aprendices_faltaron.append(aprendiz_data)

    data = {
        'lugar': encuentro.lugar,
        'fecha': encuentro.fecha.strftime('%Y-%m-%d'),
        'participantes': total_aparticipantes,
        'aprendicesAsistieron': aprendices_asistieron,
        'aprendicesFaltaron': aprendices_faltaron
    }

    return JsonResponse({'status': 'success', 'message': 'Datos retornados con exito', 'data': data}, status=200)

###############################################################################################################
#        VISTAS INFORMES
###############################################################################################################

@login_required
def informe_usuarios_x_rol(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    ws.append(["Nombre", "Apellido", "Tipo DNI", "DNI", "Ultimo ingreso", "Rol"])

    perfiles = T_perfil.objects.all()

    for perfil in perfiles:  
        last_login = perfil.user.last_login
        if last_login:  
            last_login = last_login.replace(tzinfo=None)  # quitar timezone
        else:
            last_login = ""

        ws.append([
            perfil.nom,
            perfil.apelli,
            perfil.tipo_dni,
            perfil.dni,
            last_login,
            perfil.rol
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="usuarios_rol.xlsx"'

    wb.save(response)
    return response

@login_required
def informe_fichas_x_instructor(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fichas"

    ws.append(["Ficha", "Grupo", "Institucion", "Dane Institucion",
              "Municipio", "Departamento", "Nombre", "Apellido", "Cedula"])

    fichas = T_ficha.objects.all()

    for ficha in fichas:
        ws.append([
            ficha.num,
            ficha.grupo.id,
            ficha.insti.nom,
            ficha.insti.dane,
            ficha.insti.muni.nom_munici,
            ficha.insti.muni.nom_departa.nom_departa,
            ficha.instru.perfil.nom if ficha.instru else "Sin instructor",
            ficha.instru.perfil.apelli if ficha.instru else "Sin instructor",
            ficha.instru.perfil.dni if ficha.instru else "Sin instructor",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="fichas_instructor.xlsx"'

    wb.save(response)
    return response

@login_required
def informe_fichas_x_aprendiz(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fichas"

    ws.append(["Id", "Nombre", "Apellido", "DNI", "ficha"])

    aprendices = T_apre.objects.all()

    for a in aprendices:
        ws.append([
            a.id ,
            a.perfil.nom,
            a.perfil.apelli,
            a.perfil.dni,
            a.ficha.num if a.ficha else "Sin registro"
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="fichas_aprendiz.xlsx"'

    wb.save(response)
    return response

@login_required
def informe_documentos_x_instructor_ficha(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Documentos"
    
    ws.append(["Ficha", "Grupo", "Nombre", "Apelli", "DNI", "Total documentos"])
    
    qs = (
      T_DocumentFolder.objects.filter(tipo="documento")
      .values("ficha__num", "ficha__grupo__id")
      .annotate(
          nombre = Coalesce("ficha__instru__perfil__nom", Value("Sin registro")),
          apellido = Coalesce("ficha__instru__perfil__apelli", Value("Sin registro")),
          dni = Coalesce("ficha__instru__perfil__dni", Value(None)),
          total_documentos=Count("id")
      )
      .order_by("ficha__instru__id")
    )
    
    for i in qs:
      ws.append([
        i["ficha__num"],
        i["ficha__grupo__id"],
        i["nombre"],
        i["apellido"],
        i["dni"],
        i["total_documentos"]
      ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="documentos_instructor_ficha.xlsx"'

    wb.save(response)
    return response
  
@login_required
def informe_documentos_x_instructor_aprendiz(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Documentos Aprendices"
    
    ws.append(["Ficha", "Grupo", "Nombre Instructor", "Apellido", "DNI", "Total documentos aprendices"])
    
    qs = (
        T_DocumentFolderAprendiz.objects.filter(tipo="documento")
        .values("aprendiz__ficha__num", "aprendiz__ficha__grupo__id")
        .annotate(
            nombre=Coalesce("aprendiz__ficha__instru__perfil__nom", Value("Sin registro")),
            apellido=Coalesce("aprendiz__ficha__instru__perfil__apelli", Value("Sin registro")),
            dni=Coalesce("aprendiz__ficha__instru__perfil__dni", Value(None)),
            total_documentos=Count("id")
        )
        .order_by("aprendiz__ficha__instru__id")
    )
    
    for i in qs:
        ws.append([
            i["aprendiz__ficha__num"],
            i["aprendiz__ficha__grupo__id"],
            i["nombre"],
            i["apellido"],
            i["dni"],
            i["total_documentos"]
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="documentos_instructor_aprendiz.xlsx"'

    wb.save(response)
    return response

